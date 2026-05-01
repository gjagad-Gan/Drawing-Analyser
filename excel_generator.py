"""
excel_generator.py
Produces an ISC Contract Review Sheet (.xlsx) that exactly matches the
CRS_F91TOUBR.xlsx template: 2 sheets, 31 numbered items, correct colours,
merged-cell layout, row heights and column widths.

Indo Shell Cast Private Limited — Marketing Department
"""

from __future__ import annotations

import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# Brand palette (exact hex from the template)
# ─────────────────────────────────────────────────────────────────────────────
C_NAVY      = "1E3A5F"   # header band
C_BLUE      = "2B5C99"   # section/column-header band
C_LBLUE     = "D6E4F7"   # description label cells
C_SPEC      = "FFFFFF"   # spec data cells (white)
C_REVIEW    = "EEF4FC"   # feasibility review cells
C_ALT       = "F4F6F8"   # alternating row in cost section
C_YELLOW    = "F59E0B"   # risk analysis header
C_YELROW    = "FEF3C7"   # risk label cells
C_YELDATA   = "FFF1F0"   # risk data cells (very light)
C_GREEN     = "F0FDF4"   # concluding remarks
C_META      = "D6E4F7"   # meta text colour on dark bg (actually used as fg below)
C_WHITE     = "FFFFFF"
C_DARK      = "0F172A"
C_MID       = "374151"
C_ISC_TEXT  = "1E3A5F"   # ISC-blue text on light bg


def _fill(hex6: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex6)


def _font(bold=False, color=C_DARK, size=8, name="Calibri", italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, name=name, italic=italic)


def _thin_border() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h="left", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Page 1"
    ws2 = wb.create_sheet("Page 2")

    _build_page1(ws1, data)
    _build_page2(ws2, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Low-level cell helpers
# ─────────────────────────────────────────────────────────────────────────────

def _c(ws, row: int, col: int,
       value="", bold=False, bg=None, fg=C_DARK,
       size=8, halign="left", valign="center",
       wrap=True, italic=False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, color=fg, size=size, italic=italic)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align(h=halign, v=valign, wrap=wrap)
    cell.border = _thin_border()


def _merge(ws, r1, c1, r2, c2) -> None:
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)


def _row_h(ws, row: int, h: float) -> None:
    ws.row_dimensions[row].height = h


def _col_w(ws, col: int, w: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = w


# ─────────────────────────────────────────────────────────────────────────────
# Shared header (identical on both pages)
# ─────────────────────────────────────────────────────────────────────────────

def _write_header(ws, page_of: str, data: dict) -> None:
    """Rows 1–5."""
    today = date.today().strftime("%d-%b-%Y")
    review_no = data.get("review_no") or ""

    # ── Row 1–2: Logo / brand ────────────────────────────────────────────────
    _merge(ws, 1, 1, 2, 4)
    _c(ws, 1, 1,
       "Indo Shell Cast Private Limited\nCoimbatore – 641 021  INDIA",
       bold=True, bg=C_NAVY, fg=C_WHITE, size=9, halign="center")

    _merge(ws, 1, 5, 2, 10)
    _c(ws, 1, 5,
       "MARKETING DEPARTMENT",
       bold=True, bg=C_NAVY, fg=C_WHITE, size=11, halign="center")

    _merge(ws, 1, 11, 1, 14)
    _c(ws, 1, 11,
       "Ref. No. : ISC-TS-MKTG-CRSF-R02",
       bg=C_NAVY, fg="D6E4F7", size=8, halign="left")

    _merge(ws, 2, 11, 2, 14)
    _c(ws, 2, 11,
       f"Review No. : {review_no}",
       bg=C_NAVY, fg="D6E4F7", size=8, halign="left")

    # ── Row 3–4: Title ───────────────────────────────────────────────────────
    _merge(ws, 3, 5, 4, 10)
    _c(ws, 3, 5,
       "CONTRACT REVIEW SHEET FOR ENQUIRY – RAW CASTING",
       bold=True, bg=C_NAVY, fg=C_WHITE, size=12, halign="center")

    _merge(ws, 3, 11, 3, 14)
    _c(ws, 3, 11,
       f"Date : {today}",
       bg=C_NAVY, fg="D6E4F7", size=8, halign="left")

    _merge(ws, 4, 11, 4, 14)
    _c(ws, 4, 11,
       f"Page No. :  {page_of}",
       bg=C_NAVY, fg="D6E4F7", size=8, halign="left")

    # top-left cells on rows 3–4 need fill
    for r in (3, 4):
        _merge(ws, r, 1, r, 4)
        _c(ws, r, 1, "", bg=C_NAVY)

    for r in (1, 2, 3, 4):
        _row_h(ws, r, 20 if r <= 2 else 18)

    # ── Row 5: Section title ──────────────────────────────────────────────────
    _merge(ws, 5, 1, 5, 14)
    _c(ws, 5, 1,
       "TECHNICAL REVIEW BY ENGINEERING DEPARTMENT",
       bold=True, bg=C_BLUE, fg=C_WHITE, size=9, halign="center")
    _row_h(ws, 5, 16)


# ─────────────────────────────────────────────────────────────────────────────
# Page 1
# ─────────────────────────────────────────────────────────────────────────────

def _build_page1(ws, data: dict) -> None:
    _set_col_widths_p1(ws)
    _write_header(ws, "01 of 02", data)

    # ── Row 6: Customer ───────────────────────────────────────────────────────
    _merge(ws, 6, 1, 6, 4)
    _c(ws, 6, 1, "Name of the Customer :",
       bold=True, bg=C_LBLUE, fg=C_DARK, size=8)
    _merge(ws, 6, 5, 6, 14)
    _c(ws, 6, 5, data.get("customer_name", ""),
       bold=True, bg=C_SPEC, fg=C_ISC_TEXT, size=9)
    _row_h(ws, 6, 18)

    # ── Row 7: Enquiry ref ────────────────────────────────────────────────────
    _merge(ws, 7, 1, 7, 14)
    _c(ws, 7, 1,
       f"Enquiry Ref. & Date :   {data.get('enquiry_ref', '')}",
       bg="F4F6F8", fg=C_DARK, size=8)
    _row_h(ws, 7, 18)

    # ── Rows 8–9: End use / Source ────────────────────────────────────────────
    _merge(ws, 8, 1, 9, 7)
    _c(ws, 8, 1,
       f"End use of the product :\n{data.get('end_use', '')}",
       bg=C_LBLUE, fg=C_DARK, size=8)
    _merge(ws, 8, 8, 9, 14)
    _c(ws, 8, 8,
       f"Source of Information :\n{data.get('source_of_information', '')}",
       bg=C_LBLUE, fg=C_DARK, size=8)
    for r in (8, 9):
        _row_h(ws, r, 20)

    # ── Rows 10–11: Column headers ────────────────────────────────────────────
    _merge(ws, 10, 1, 11, 1)
    _c(ws, 10, 1, "S.\nNo.", bold=True, bg=C_BLUE, fg=C_WHITE,
       size=8, halign="center")
    _merge(ws, 10, 2, 11, 4)
    _c(ws, 10, 2, "DESCRIPTION", bold=True, bg=C_BLUE, fg=C_WHITE,
       size=8, halign="center")
    _merge(ws, 10, 5, 11, 8)
    _c(ws, 10, 5, "SPECIFIED IN ENQUIRY / DRAWING / SPECIFICATION",
       bold=True, bg=C_BLUE, fg=C_WHITE, size=8, halign="center")
    _merge(ws, 10, 9, 11, 14)
    _c(ws, 10, 9, "FEASIBILITY REVIEW",
       bold=True, bg=C_BLUE, fg=C_WHITE, size=8, halign="center")
    for r in (10, 11):
        _row_h(ws, r, 22)

    # ── Items 1–11 ────────────────────────────────────────────────────────────
    items = data.get("page1", [])
    row = 12
    for idx, item in enumerate(items[:11]):
        # Span 3, 4 or 6 rows depending on text volume in spec column
        spec_lines = len((item.get("spec") or "").split("\n"))
        if spec_lines <= 3:
            n_rows, row_h = 3, 42
        elif spec_lines <= 6:
            n_rows, row_h = 4, 42
        else:
            n_rows, row_h = 6, 42
        _write_item_p1(ws, row, item, n_rows=n_rows, row_h=row_h)
        row += n_rows

    # ── Sign-off ──────────────────────────────────────────────────────────────
    _merge(ws, row, 1, row + 1, 8)
    _c(ws, row, 1, "Prepared by :", bold=True, bg=C_LBLUE, fg=C_ISC_TEXT, size=8)
    _merge(ws, row, 9, row + 1, 14)
    _c(ws, row, 9, "Approved by :", bold=True, bg=C_LBLUE, fg=C_ISC_TEXT, size=8)
    for r in (row, row + 1):
        _row_h(ws, r, 18)
    row += 2

    _merge(ws, row, 1, row + 1, 8)
    _c(ws, row, 1, "Date :", bg=C_SPEC, fg=C_MID, size=8)
    _merge(ws, row, 9, row + 1, 14)
    _c(ws, row, 9, "Date :", bg=C_SPEC, fg=C_MID, size=8)
    for r in (row, row + 1):
        _row_h(ws, r, 18)


def _write_item_p1(ws, start_row: int, item: dict,
                   n_rows: int = 3, row_h: float = 42) -> None:
    end_row = start_row + n_rows - 1

    _merge(ws, start_row, 1, end_row, 1)
    _c(ws, start_row, 1, item.get("no", ""), bold=True,
       bg=C_LBLUE, fg=C_ISC_TEXT, halign="center")

    _merge(ws, start_row, 2, end_row, 4)
    _c(ws, start_row, 2, item.get("description", ""),
       bold=True, bg=C_LBLUE, fg=C_DARK)

    _merge(ws, start_row, 5, end_row, 8)
    _c(ws, start_row, 5, item.get("spec", ""), bg=C_SPEC, fg=C_DARK)

    _merge(ws, start_row, 9, end_row, 14)
    _c(ws, start_row, 9, item.get("review", ""), bg=C_REVIEW, fg=C_MID)

    for r in range(start_row, end_row + 1):
        ws.row_dimensions[r].height = row_h


def _set_col_widths_p1(ws) -> None:
    widths = {
        1: 4.5,    # A
        2: 8.0,    # B
        3: 14.0,   # C
        4: 10.33,  # D
        5: 5.0,    # E
        6: 9.0,    # F
        7: 9.0,    # G
        8: 12.16,  # H
        9: 5.5,    # I
        10: 8.0,   # J
        11: 8.0,   # K
        12: 8.0,   # L
        13: 8.0,   # M
        14: 8.0,   # N
    }
    for col, w in widths.items():
        _col_w(ws, col, w)


# ─────────────────────────────────────────────────────────────────────────────
# Page 2
# ─────────────────────────────────────────────────────────────────────────────

def _build_page2(ws, data: dict) -> None:
    _set_col_widths_p2(ws)
    _write_header(ws, "02 of 02", data)

    # ── Rows 6–7: Column headers (cost section) ───────────────────────────────
    _merge(ws, 6, 1, 7, 1)
    _c(ws, 6, 1, "S.\nNo.", bold=True, bg=C_BLUE, fg=C_WHITE,
       halign="center")
    _merge(ws, 6, 2, 7, 4)
    _c(ws, 6, 2, "DESCRIPTION", bold=True, bg=C_BLUE, fg=C_WHITE,
       halign="center")
    _merge(ws, 6, 5, 7, 14)
    _c(ws, 6, 5, "INPUTS FOR COST ESTIMATION",
       bold=True, bg=C_BLUE, fg=C_WHITE, halign="center")
    for r in (6, 7):
        _row_h(ws, r, 22)

    # ── Items 12–25 (cost estimation) ────────────────────────────────────────
    cost_items = data.get("page2_cost", [])
    row = 8
    for i, item in enumerate(cost_items[:14]):
        bg = C_SPEC if i % 2 == 0 else C_ALT
        _merge(ws, row, 1, row, 1)
        _c(ws, row, 1, item.get("no", ""), bold=True, bg=C_LBLUE, fg=C_ISC_TEXT,
           halign="center")
        _merge(ws, row, 2, row, 4)
        _c(ws, row, 2, item.get("description", ""), bold=True, bg=C_LBLUE,
           fg=C_DARK)
        _merge(ws, row, 5, row, 14)
        _c(ws, row, 5, item.get("value", ""), bg=bg, fg=C_DARK)
        _row_h(ws, row, 22)
        row += 1

    # ── Spacer row ────────────────────────────────────────────────────────────
    _row_h(ws, row, 10)
    row += 1

    # ── Risk analysis header ──────────────────────────────────────────────────
    _merge(ws, row, 1, row, 14)
    _c(ws, row, 1, "RISK ANALYSIS", bold=True, bg=C_YELLOW, fg=C_DARK,
       halign="center")
    _row_h(ws, row, 16)
    row += 1

    # ── Items 26–31 (risk analysis) ───────────────────────────────────────────
    risk_items = data.get("page2_risk", [])
    # Each risk item spans 3 rows (28pt each, except item 29 which is taller)
    risk_heights = [
        (28, 28, 28),  # 26
        (28, 28, 28),  # 27
        (28, 28, 28),  # 28
        (62, 28, 28),  # 29 — taller for mitigation detail
        (28, 28, 28),  # 30
        (28, 28, 28),  # 31
    ]
    for idx, item in enumerate(risk_items[:6]):
        h = risk_heights[idx] if idx < len(risk_heights) else (28, 28, 28)
        _write_risk_item(ws, row, item, h)
        row += 3

    # ── Spacer ────────────────────────────────────────────────────────────────
    _row_h(ws, row, 12)
    row += 1

    # ── Concluding remarks ────────────────────────────────────────────────────
    _merge(ws, row, 1, row + 3, 4)
    _c(ws, row, 1, "Concluding Remarks :", bold=True, bg=C_GREEN, fg=C_ISC_TEXT)
    _merge(ws, row, 5, row + 3, 14)
    _c(ws, row, 5, data.get("concluding_remarks", ""), bg=C_GREEN, fg=C_DARK)
    for r in range(row, row + 4):
        _row_h(ws, r, 22)


def _write_risk_item(ws, start_row: int, item: dict, heights: tuple) -> None:
    r1, r2, r3 = start_row, start_row + 1, start_row + 2
    _merge(ws, r1, 1, r3, 1)
    _c(ws, r1, 1, item.get("no", ""), bold=True, bg=C_YELROW,
       fg=C_DARK, halign="center")
    _merge(ws, r1, 2, r3, 4)
    _c(ws, r1, 2, item.get("description", ""), bold=True, bg=C_YELROW, fg=C_DARK)
    _merge(ws, r1, 5, r3, 14)
    _c(ws, r1, 5, item.get("value", ""), bg=C_YELDATA, fg=C_DARK)
    for r, h in zip((r1, r2, r3), heights):
        ws.row_dimensions[r].height = h


def _set_col_widths_p2(ws) -> None:
    widths = {
        1: 4.5,    # A
        2: 8.0,    # B
        3: 14.0,   # C
        4: 4.5,    # D
        5: 7.5,    # E
        6: 8.0, 7: 8.0, 8: 8.0, 9: 6.0,
        10: 8.0, 11: 8.0, 12: 8.0, 13: 8.0, 14: 8.0,
    }
    for col, w in widths.items():
        _col_w(ws, col, w)
